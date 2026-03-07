import pandas as pd
from datetime import datetime
import json
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import xlwings as xw
import shutil
import openpyxl
from docx import Document

excel_path='1.xlsx'
word_path='1.docx'

wb=openpyxl.load_workbook(excel_path,data_only=True)
sheet1=wb.worksheets[0]
sheet2=wb.worksheets[1]

doc=Document(word_path)
tables=doc.tables

if len(tables)<2:
    raise Exception('Word文档中至少需要2个表格')
table1=tables[0]
table2=tables[1]

def fill_table_from_sheet(table,sheet):
    written_cells=set()
    for i , row in enumerate(table.rows):
        for j , cell in enumerate(row.cells):
            if cell._tc in written_cells:
                continue
            written_cells.add(cell._tc)
            excel_value=sheet.cell(row=i+1,column=j+1).value
            cell.text=str(excel_value) if excel_value is not None else ''
            # if excel_value is not None:
            #     cell.text=str(excel_value)
            # else:
            #     cell.text=''

fill_table_from_sheet(table1,sheet1)
fill_table_from_sheet(table2,sheet2)

doc.save('1_填充后.docx')
print('数据填充完成')